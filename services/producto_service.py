# services/producto_service.py
from models import Producto
from repositories.producto_repository import ProductoRepository

class ProductoService:
    # Configuración de precio recomendado
    PORCENTAJE_RECOMENDADO = 80  # +80% de aumento
    
    def __init__(self, repository: ProductoRepository):
        self._repository = repository

    def get_all_products(self):
        return self._repository.get_all()

    def get_precio_recomendado(self, precio):
        """
        Calcula el precio recomendado (+80%) para un producto.
        
        Args:
            precio: Precio unitario del producto
        
        Returns:
            float: Precio recomendado (precio * 1.80)
        """
        precio_recomendado = precio * (1 + self.PORCENTAJE_RECOMENDADO / 100)
        return round(precio_recomendado, 2)

    def create_product(self, codigo, descripcion, precio):
        if self._repository.get_by_code(codigo):
            raise ValueError(f"Producto con código {codigo} ya existe.")
        
        nuevo_producto = Producto(codigo, descripcion, precio)
        self._repository.add(nuevo_producto)
        return nuevo_producto

    def update_product(self, codigo, descripcion, precio):
        return self._repository.update(codigo, descripcion, precio)

    def delete_product(self, codigo):
        # Future: check if product is in any non-finalized invoice, etc.
        return self._repository.delete(codigo)

    def apply_price_increase(self, producto_codigos, porcentaje):
        """
        Aplica un cambio de precio a múltiples productos.
        
        Args:
            producto_codigos: Lista de códigos de productos
            porcentaje: Porcentaje de cambio (ej: 10 para +10%, -5 para -5%)
        
        Returns:
            Número de productos actualizados
        
        Raises:
            ValueError: Si porcentaje == 0, porcentaje < -100 o si no hay productos a actualizar
        """
        if porcentaje == 0:
            raise ValueError("El porcentaje no puede ser 0")
        
        if porcentaje < -100:
            raise ValueError("El porcentaje no puede ser menor a -100%")
        
        if not producto_codigos or len(producto_codigos) == 0:
            raise ValueError("Debe seleccionar al menos un producto")
        
        actualizados = 0
        for codigo in producto_codigos:
            producto = self._repository.get_by_code(codigo)
            if producto:
                # Calcular nuevo precio
                nuevo_precio = producto.precio * (1 + porcentaje / 100)
                # Redondear a 2 decimales
                nuevo_precio = round(nuevo_precio, 2)
                # Actualizar
                self._repository.update(codigo, producto.descripcion, nuevo_precio)
                actualizados += 1
        
        return actualizados

    def export_to_excel(self, filepath):
        """
        Exporta todos los productos a un archivo Excel (.xlsx).
        
        Args:
            filepath: Ruta completa del archivo Excel a crear
        
        Returns:
            True si la exportación fue exitosa
        
        Raises:
            ValueError: Si no hay productos para exportar
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, numbers
        
        productos = self.get_all_products()
        
        if not productos:
            raise ValueError("No hay productos para exportar")
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Productos"
        
        # Crear headers
        headers = ["Código", "Descripción", "Precio", "Precio Recomendado (+80%)"]
        ws.append(headers)
        
        # Aplicar negrita a headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        # Agregar datos de productos
        for producto in productos:
            precio_recomendado = self.get_precio_recomendado(producto.precio)
            ws.append([
                producto.codigo,
                producto.descripcion,
                producto.precio,
                precio_recomendado
            ])
        
        # Aplicar formato de moneda a columnas de precio
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=4):
            for cell in row:
                cell.number_format = '$#,##0.00'
        
        # Guardar archivo
        wb.save(filepath)
        return True
